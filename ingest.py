# ingest.py
# Usage: py ingest.py ./docs
#        py ingest.py ./docs --clear   (wipe + re-ingest)

import os, sys, json, base64
from pathlib import Path
from dotenv import load_dotenv
import openai
from supabase import create_client

# Lazy-import unstructured (heavy, only needed for PDF/DOCX/PPTX)
partition = None
chunk_by_title = None

def _load_unstructured():
    global partition, chunk_by_title
    if partition is None:
        print('  Loading unstructured library (first time may be slow)...')
        from unstructured.partition.auto import partition as _p
        from unstructured.chunking.title import chunk_by_title as _c
        partition = _p
        chunk_by_title = _c

load_dotenv()

# --- Provider config (Ollama local by default, OpenAI if OPENAI_API_KEY is set) ---
if os.environ.get('OPENAI_API_KEY'):
    ai = openai.OpenAI(api_key=os.environ['OPENAI_API_KEY'])
    EMBED_MODEL  = os.environ.get('EMBED_MODEL', 'text-embedding-3-small')
    VISION_MODEL = os.environ.get('CHAT_MODEL',  'gpt-4o-mini')
    print('Using OpenAI cloud')
else:
    ai = openai.OpenAI(
        api_key='ollama',
        base_url=os.environ.get('OLLAMA_BASE_URL', 'http://localhost:11434/v1'),
    )
    EMBED_MODEL  = os.environ.get('EMBED_MODEL', 'nomic-embed-text')
    VISION_MODEL = os.environ.get('CHAT_MODEL',  'llama3.2:3b')
    print('Using Ollama local')

supabase = create_client(os.environ['SUPABASE_URL'],
                         os.environ['SUPABASE_SERVICE_KEY'])

SUPPORTED  = {'.pdf', '.docx', '.pptx', '.xlsx', '.txt', '.md', '.html', '.htm'}
SIMPLE_TEXT = {'.txt', '.md', '.html', '.htm'}  # can be chunked without unstructured
IMAGE_EXT  = {'.png', '.jpg', '.jpeg', '.webp', '.tiff'}
BATCH_SIZE = 64   # max texts per embedding API call


def simple_chunk(text: str, max_chars: int = 1500, overlap: int = 200) -> list[str]:
    """Split plain text into overlapping chunks by paragraph boundaries."""
    paragraphs = [p.strip() for p in text.split('\n\n') if p.strip()]
    # Sub-split any paragraph that's longer than max_chars
    split_paras = []
    for para in paragraphs:
        if len(para) <= max_chars:
            split_paras.append(para)
        else:
            # Split on single newlines first, then hard-split if needed
            lines = para.split('\n')
            sub = ''
            for line in lines:
                if len(sub) + len(line) + 1 > max_chars and sub:
                    split_paras.append(sub)
                    sub = line
                else:
                    sub = sub + '\n' + line if sub else line
            if sub:
                # Hard-split any remaining chunk that's still too long
                while len(sub) > max_chars:
                    split_paras.append(sub[:max_chars])
                    sub = sub[max_chars - overlap:]
                split_paras.append(sub)

    chunks = []
    current = ''
    for para in split_paras:
        if len(current) + len(para) + 2 > max_chars and current:
            chunks.append(current)
            # keep overlap from end of previous chunk
            current = current[-overlap:] + '\n\n' + para if overlap else para
        else:
            current = current + '\n\n' + para if current else para
    if current and len(current) > 40:
        chunks.append(current)
    return chunks


def embed(texts: list[str]) -> list[list[float]]:
    """Embed a list of strings, respecting batch limits."""
    results = []
    for i in range(0, len(texts), BATCH_SIZE):
        batch = texts[i : i + BATCH_SIZE]
        resp  = ai.embeddings.create(model=EMBED_MODEL, input=batch)
        results.extend([d.embedding for d in resp.data])
    return results


def extract_metadata(chunk) -> dict:
    """Pull per-chunk metadata from unstructured element metadata."""
    meta = getattr(chunk, 'metadata', None)
    if not meta:
        return {}
    return {k: v for k, v in {
        'page_number':   getattr(meta, 'page_number',  None),
        'section_title': str(getattr(meta, 'section', '') or ''),
        'slide_number':  getattr(meta, 'page_number',  None),  # PPTX reuses page_number
        'sheet_name':    getattr(meta, 'sheet_name',   None),
    }.items() if v}


def ingest_document(path: Path, clear_existing: bool = False) -> int:
    """Parse, chunk, embed, and upsert one document. Returns chunk count."""
    suffix = path.suffix.lower()

    if suffix not in SUPPORTED:
        print(f'  SKIP (unsupported): {path.name}')
        return 0

    if clear_existing:
        supabase.table('chunks').delete().eq('source', path.name).execute()

    print(f'  Parsing {path.name} ...', end='', flush=True)

    # --- XLSX: special handling — row-based chunking ---
    if suffix == '.xlsx':
        return ingest_xlsx(path)

    # --- Plain text / Markdown: fast path (no unstructured needed) ---
    if suffix in SIMPLE_TEXT:
        raw = path.read_text(encoding='utf-8', errors='replace')
        texts = simple_chunk(raw)
    else:
        # --- PDF, DOCX, PPTX: use unstructured ---
        _load_unstructured()
        elements = partition(
            filename=str(path),
            strategy='hi_res',
            extract_images_in_pdf=False,
            languages=['eng'],
        )
        chunks = chunk_by_title(
            elements,
            max_characters=1500,
            overlap=200,
            include_orig_elements=False,
        )
        texts = [str(c).strip() for c in chunks]

    texts = [t for t in texts if len(t) > 40]  # discard near-empty chunks

    if not texts:
        print(' (no extractable text)')
        return 0

    embeddings = embed(texts)

    rows = [
        {
            'source':    path.name,
            'file_type': suffix,
            'content':   texts[i],
            'embedding': embeddings[i],
            'metadata':  {},
        }
        for i in range(len(texts))
    ]

    supabase.table('chunks').insert(rows).execute()
    print(f' {len(rows)} chunks')
    return len(rows)


def ingest_xlsx(path: Path) -> int:
    """
    Excel: describe each sheet as a natural-language summary.
    Raw cell dumps are terrible for semantic retrieval.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_out = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value) for c in next(ws.iter_rows(max_row=1))
                   if c.value is not None]
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                data_rows.append(dict(zip(headers, row)))
            if len(data_rows) >= 50:   # cap for PoC
                break

        if not data_rows:
            continue

        # Ask the LLM to describe the sheet in natural language
        sample_json = json.dumps(data_rows[:10], default=str, indent=2)
        description = ai.chat.completions.create(
            model=VISION_MODEL,
            messages=[{
                'role': 'user',
                'content': (
                    f'Describe this spreadsheet sheet ({sheet_name}) for a search index. '
                    f'Include column names, data types, and what the data represents.\n\n'
                    f'Sample data (first 10 rows):\n{sample_json}'
                )
            }]
        ).choices[0].message.content

        text = f'Sheet: {sheet_name}\n{description}'
        rows_out.append({
            'source':    path.name,
            'file_type': '.xlsx',
            'content':   text,
            'embedding': embed([text])[0],
            'metadata':  {'sheet_name': sheet_name, 'row_count': len(data_rows)},
        })

    if rows_out:
        supabase.table('chunks').insert(rows_out).execute()
    print(f' {len(rows_out)} sheet chunks')
    return len(rows_out)


def describe_image(path: Path) -> str:
    """Use vision model to produce a searchable text description of an image."""
    with open(path, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode()
    ext  = path.suffix.lstrip('.').lower()
    mime = {'jpg': 'jpeg', 'tiff': 'tiff'}.get(ext, ext)
    return ai.chat.completions.create(
        model=VISION_MODEL,
        messages=[{'role': 'user', 'content': [
            {'type': 'image_url',
             'image_url': {'url': f'data:image/{mime};base64,{b64}'}},
            {'type': 'text',
             'text': (
                'Describe this image in detail for a document search index. '
                'Include all visible text, chart titles, axis labels, '
                'table headers, diagram labels, and key visual elements. '
                'Be thorough — this description is the only way this image '
                'will be found by search.'
             )}
        ]}]
    ).choices[0].message.content


def ingest_image(path: Path) -> int:
    description = describe_image(path)
    row = {
        'source':    path.name,
        'file_type': 'image',
        'content':   description,
        'embedding': embed([description])[0],
        'metadata':  {
            'source_image':      path.name,
            'description_model': VISION_MODEL,
        }
    }
    supabase.table('chunks').insert([row]).execute()
    print(f'  Image {path.name}: described + indexed')
    return 1


if __name__ == '__main__':
    docs_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path('./docs')
    clear    = '--clear' in sys.argv

    all_files    = list(docs_dir.iterdir())
    total_chunks = 0
    skipped      = 0

    print(f'Ingesting {len(all_files)} files from {docs_dir}...')

    for f in sorted(all_files):
        if f.suffix.lower() in IMAGE_EXT:
            total_chunks += ingest_image(f)
        elif f.suffix.lower() in SUPPORTED:
            total_chunks += ingest_document(f, clear_existing=clear)
        else:
            skipped += 1

    print(f'\nDone. {total_chunks} total chunks indexed. {skipped} files skipped.')
